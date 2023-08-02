#Improtar librerias necesarias
from typing import List, Tuple, Dict, Optional, Union
from openpyxl.styles import Font, Alignment
from itertools import combinations
from itertools import permutations
import matplotlib.pyplot as plt
from gurobipy import *
import pandas as pd
import numpy as np
from pulp import *
import openpyxl
import os

class SecuenciacionProgramacionLineal():
    # Inicializar variables
    def __init__(self, tareas:dict):
        '''
        SecuenciacionProgramacionLineal(tareas: dict[str ,dict[float, ..., float]])\n
        tareas: diccionario de key:tareas value:diccionarios de maquinas con duraciones 
        de cada tarea respectiva en la maquina, en caso de estar repetida la maquina en
        el proceso se debe agregar el nombre de la maquinaseguido de raya al piso y el 
        numero de la priorida, se debe ingresar en el diccionario de cada tarea las 
        maquinas en el orden que van a ser procesados
        Ejemplo:\n
        tareas={'T1' : { 'M1':5  , 'M2':7   , 'M3':7   },\n
                'T2' : { 'M1_1':3 , 'M2_1':7 , 'M1_2':2  , 'M2_2':7 },\n
                'T3' : { 'M3_1':5 , 'M2':8   , 'M3_2':9 },\n
                'T4' : { 'M1':4   , 'M2':7   , 'M3':6   },\n
                'T5' : { 'M2_1':5 , 'M1':6   , 'M2_2':7  , 'M3':2 },\n
                'T6' : { 'M1':8   , 'M3':5   , 'M2':4   },\n
                'T7' : { 'M3_1':9 , 'M2':2   , 'M1':5    , 'M3_2':5 },\n
                'T8' : { 'M2_1':3 , 'M3':4   , 'M2_2':1 }}\n
        SecuanciacionPL = SecuenciacionProgramacionLineal(tareas)
        TiemposInicioResultado = SecuanciacionPL.tiempos_resultado
        print(TiemposInicioResultado)
        SecuanciacionPL.DiagramaGantt()
        '''
        self.modelo = LpProblem("modelo_secuanciacion_maquinas", sense=LpMinimize)
        self.M = 0
        for tarea in tareas.keys():
            for maquina in tareas[tarea]:
                self.M += tareas[tarea][maquina] * 2
        self.M = int(self.M)
        self.tareas = tareas
        self.CrearVariables()
        self.RestriccionesSolapeTiempos()
        self.RestriccionSecuencia()
        self.RestriccionesTiempoMaximo()
        self.FuncionObjetivo()
        self.Solucionar()
        self.DiccionarioTiemposInicio()

    # Crear vairbales modelo
    def CrearVariables(self):
        self.nombres_tareas = list(self.tareas.keys())
        self.nombres_maquinas = list(set([maquina.split('_')[0] for tarea in self.tareas.keys() for maquina in self.tareas[tarea]]))
        #self.nombres_maquinas_tareas = dict([(tarea,list(self.tareas[tarea].keys())) for tarea in self.nombres_tareas]) CAMBIO POSIBLE
        self.nombres_maquinas_tareas = {tarea:list(self.tareas[tarea].keys()) for tarea in self.nombres_tareas}
        # Crear variables Tiempos de inicio
        self.nombres_tiempos_inicio = [tarea+'_'+maquina for tarea in self.nombres_tareas for maquina in self.nombres_maquinas_tareas[tarea]]
        self.tiempos_inicio = LpVariable.dicts("TiemposInicio", self.nombres_tiempos_inicio , lowBound=0, cat='Continuos')
        # Crear agrupacion por Tareas
        self.diccionario_nombres_secuencia = dict((tarea,[]) for tarea in self.nombres_tareas)
        for nombres_tiempos_inicio in self.nombres_tiempos_inicio:
            nombres_tiempos_inicio_split = nombres_tiempos_inicio.split('_')
            self.diccionario_nombres_secuencia[nombres_tiempos_inicio_split[0]].append(nombres_tiempos_inicio)
        # Crear agrupacion por Maquinas
        self.diccionario_nombres_maquinas = dict((maquina,[]) for maquina in self.nombres_maquinas)
        for maquina in self.nombres_maquinas:
            for nombres_tiempos_inicio in self.nombres_tiempos_inicio:
                if maquina in nombres_tiempos_inicio:
                    self.diccionario_nombres_maquinas[maquina].append(nombres_tiempos_inicio)
        # Crear variables Binarias de activacion
        self.nombres_binarias = []
        for maquina in self.nombres_maquinas:
            self.nombres_binarias += list(combinations(self.diccionario_nombres_maquinas[maquina],2))
        self.binaria_activacion = LpVariable.dicts("BinariaActivacion", self.nombres_binarias , cat='Binary')
    
    # Restriccion de secuenciacion
    def RestriccionesSolapeTiempos(self):
        for nombres_tiempos_inicio in self.nombres_binarias:
            nti1, nti2 = nombres_tiempos_inicio[0], nombres_tiempos_inicio[1] 
            tnti1, tnti2 = nti1.split('_',1)[0], nti2.split('_',1)[0]
            mnti1, mnti2 = nti1.split('_',1)[1], nti2.split('_',1)[1]
            dur1, dur2 = self.tareas[tnti1][mnti1], self.tareas[tnti2][mnti2]
            self.modelo += self.tiempos_inicio[nti1] + dur1 <= self.tiempos_inicio[nti2] + self.binaria_activacion[nti1, nti2]*self.M
            self.modelo += self.tiempos_inicio[nti2] + dur2 <= self.tiempos_inicio[nti1] + (1-self.binaria_activacion[nti1, nti2])*self.M

    # Restriccion de Tiempo Maximo
    def RestriccionesTiempoMaximo(self):
        self.TiempoMinimo = LpVariable("TiemposMinimo", lowBound=0, cat='Continuos')
        for tarea in self.diccionario_nombres_secuencia.keys():
            ultimaMaquina = self.diccionario_nombres_secuencia[tarea][-1]
            tnti, mnti = ultimaMaquina.split('_',1)[0], ultimaMaquina.split('_',1)[1] 
            self.modelo += self.tiempos_inicio[ultimaMaquina] + self.tareas[tnti][mnti] <= self.TiempoMinimo
    
    # Restriccion de Secuencia
    def RestriccionSecuencia(self):
        for tarea in self.diccionario_nombres_secuencia.keys():
            for n in range(len(self.diccionario_nombres_secuencia[tarea])-1):
                nti1, nti2 = self.diccionario_nombres_secuencia[tarea][n], self.diccionario_nombres_secuencia[tarea][n+1]
                tnti1 = nti1.split('_',1)[0]
                mnti1 = nti1.split('_',1)[1]
                self.modelo += self.tiempos_inicio[nti1] + self.tareas[tnti1][mnti1] <= self.tiempos_inicio[nti2]
    
    # Declaracion de Funcion objetivo
    def FuncionObjetivo(self):
        self.modelo += self.TiempoMinimo

    # Solucionar modelo
    def Solucionar(self):
        self.status = self.modelo.solve( solver = GUROBI(msg = False) )
        if LpStatus[self.modelo.status] == 'Optimal':
            self.horizonteTemporal = round(value(self.modelo.objective),0)
        else:
            raise ValueError('Modelo en estado: '+LpStatus[self.modelo.status])

    # Diccionario de tiempos de inicio
    def DiccionarioTiemposInicio(self):
        self.tiempos_resultado = {}
        for v in self.modelo.variables():
            if 'TiemposInicio' in str(v):
                nombre = str(v)
                nombre = nombre.replace('TiemposInicio_','')
                self.tiempos_resultado[nombre] = round(v.varValue,0)
        return self.tiempos_resultado

    # Generar Diagrama de Gantt
    def DiagramaGantt(self):
        self.DiccionarioTiemposInicio()
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('Maquinas')
        for tareas in self.nombres_tareas:
            inicios = []
            maquinas = []
            duraciones = []
            for nombreInicio in self.tiempos_resultado.keys():
                if tareas in nombreInicio:
                    inicios.append(self.tiempos_resultado[nombreInicio])
                    tar, maq = nombreInicio.split('_',1)[0], nombreInicio.split('_',1)[1] 
                    duraciones.append(self.tareas[tar][maq])
                    maquinas.append(maq.split('_')[0])
            ax.barh(maquinas, duraciones, left=inicios, label=tareas)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show()

class BalanceoLienaProgramacionLineal():
    def __init__(self, tareas, produccionDiaraDeseada, produccionDiaraActual):
        '''
        BalanceoLienaProgramacionLineal(tareas: dict[str, list[float, str]], produccionDiaraDeseada: int, produccionDiaraActual: int)\n
        Enviar diccionario con nombre de cada tarea como llave, y cada valor debe ser una lista en la cual el primer valor debe ser
        un valor tipo int o float que representara el tiempo en SEGUNDOS de la tarea, y en la siguiente posicion de la lista un str
        con el nombre del predecesor, en caso de ser mas de un predecesor separarlos con el simbolo '-', es decir es un simbolo reservado 
        para su uso, y evitar el uso usar espacios
        Para el segundo argumento se debe enviar la capacidad de unidades que desea realizar en un dia
        Para el tercer argumento se debe enviar la capacidad de unidades que puede realizar en un dia
        Ejemplo:\n
        tareas={'A' : [12   , '-'],\n
                'B' : [24   , '-'],\n
                'C' : [42   , 'A'] ,\n
                'D' : [6    , 'A-B'],\n
                'E' : [18   , 'B'],\n
                'F' : [6.6  , 'C'],\n
                'G' : [19.2 , 'C'],\n
                'H' : [36   , 'C-D'],\n
                'I' : [16.2 , 'F-G-H'],\n
                'J' : [22.8 , 'E-H'],\n
                'K' : [30   , 'I-J'] }\n
        produccionDiaraDeseada =  500\n
        tiempoFuncionamientoDiario = 500\n
        '''
        self.modelo = LpProblem("modelo_balanceo_linea", sense=LpMinimize)
        self.tareas = tareas
        self.unidadesDeseadas = produccionDiaraDeseada
        self.unidadesActual = produccionDiaraActual
        self.CalcularTakeTime()
        self.EstacionesMinimas()
        self.CrearVariables()
        self.RestriccionPredecesores()
        self.RestriccionActivaciones()
        self.FuncionObjetivo()
        self.Solucionar()
    
    # Calcular Take time
    def CalcularTakeTime(self):
        self.takeTime = (self.unidadesActual*60)/self.unidadesDeseadas
        for tarea in self.tareas.keys():
            if self.tareas[tarea][0] >= self.takeTime:
                self.takeTime = self.tareas[tarea][0]
    
    # Calcular minimo de estaciones sugeridas, mas un margen
    def EstacionesMinimas(self):
        self.estaciones = sum([self.tareas[tarea][0] for tarea in self.tareas.keys()])/self.takeTime
        self.estaciones = int((self.estaciones//1)+1)+1 # + 1 Extra para descartar infactibilidad del modelo
        print(self.estaciones)
    
    # Crear vairbales modelo
    def CrearVariables(self):
        self.BinariaEstacion = LpVariable.dicts("BinariaEstacion", (estacion for estacion in range(self.estaciones)) , cat='Binary')
        self.BinariaTareaEstacion = LpVariable.dicts("BinariaTareaEstacion", ((tarea,estacion) for estacion in range(self.estaciones) for tarea in self.tareas.keys()) , cat='Binary')

    # Restriccion de predecesores
    def RestriccionPredecesores(self):
        for tarea in self.tareas.keys():
            if self.tareas[tarea][1] != '-':
                predecesores = self.tareas[tarea][1].split('-')
                for estacion in range(self.estaciones):
                    self.modelo += lpSum(self.BinariaTareaEstacion[predecesor,estacionacu] for predecesor in predecesores for estacionacu in range(0,estacion+1)) >= self.BinariaTareaEstacion[tarea,estacion]*len(predecesores) 

    # Restriccion Activaciones de estaciones y tareas por estacion
    def RestriccionActivaciones(self):
        for estacion in range(self.estaciones):
            self.modelo += lpSum(self.BinariaTareaEstacion[tarea,estacion]*self.tareas[tarea][0] for tarea in self.tareas.keys()) <= self.takeTime*self.BinariaEstacion[estacion]
        for tarea in self.tareas.keys():
            self.modelo += lpSum(self.BinariaTareaEstacion[tarea,estacion] for estacion in range(self.estaciones)) == 1

    # Declaracion de Funcion objetivo
    def FuncionObjetivo(self):
        self.modelo += lpSum(self.BinariaEstacion[estacion] for estacion in range(self.estaciones))

    # Solucionar modelo multiobjetivo
    def Solucionar(self):
        # Modelo Uso minimo de estaciones
        self.status = self.modelo.solve(solver=GUROBI(msg = False))
        if 'Optimal'== LpStatus[self.modelo.status]:
            self.horizonteTemporal = round(value(self.modelo.objective),0)
        else:
            raise 'Porblema en factibilidad del modelo'
        estaciones = 0
        # Asignacion de Restriccion Minima de Estaciones
        for v in self.modelo.variables():
            if 'BinariaEstacion' in str(v):
                estaciones += v.varValue
        self.maximo_tiempo_estacion = LpVariable('MaximoTiempoEstacion', lowBound=0, cat='Continuous')
        for estacion in range(self.estaciones):
            self.modelo += lpSum(self.BinariaTareaEstacion[tarea,estacion]*self.tareas[tarea][0] for tarea in self.tareas.keys()) <= self.maximo_tiempo_estacion
        self.modelo += lpSum(self.BinariaEstacion[estacion] for estacion in range(self.estaciones)) == estaciones 
        self.modelo += (1/sum([self.tareas[tarea][0] for tarea in self.tareas.keys()]))*(estaciones*self.maximo_tiempo_estacion)
        # Asignacion Maximizar Eficiencia de Linea en modelo
        self.status = self.modelo.solve(solver=GUROBI(msg = False))
        if 'Optimal'== LpStatus[self.modelo.status]:
            print('-'*5+' Modelo solucionado correctamente '+'-'*5)
            self.horizonteTemporal = round(value(self.modelo.objective),0)
        else:
            raise 'Porblema en factibilidad del modelo'
    
    # Diccionario tareas por estacion
    def Diccionario_Estaciones(self):
        self.ActivacionEstacion = {}
        for v in self.modelo.variables():
            if 'BinariaTareaEstacion' in str(v) and v.varValue>0:
                nombre = str(v)
                nombre = nombre.replace('BinariaTareaEstacion_','')
                nombre = nombre.replace('(','')
                nombre = nombre.replace(')','')
                nombre = nombre.replace('_','')
                nombre = nombre.replace("'",'')
                nombre = nombre.split(',')
                self.ActivacionEstacion[nombre[0]] = 'Estacion '+ str(int(nombre[1])+1)
        return self.ActivacionEstacion

class SecuenciacionReglaJhonson():
    '''
    SecuenciacionReglaJhonson(tareas: dict[str ,dict[str:float, str:float, str:float]])\n
    Aplicacion de la regla de Jonson, para esto se debe enviar un diccioanrio de diccioanrios
    de la siguiente forma, tareas: diccionario de key:tareas value:diccionarios de maquinas con duraciones 
    de cada tarea respectiva en la maquina, en este caso no todas las tareas deben ser procesadas en el mismo
    orden en a traves de cada una de las maquinas, esta heuristica acepta menor o igual a 3 maquinas
    Ejemplo:\n
    tareas={'T1' :{'M1':3,'M2':7,'M3':3},\n
            'T2' :{'M1':1,'M2':4,'M3':9},\n
            'T3' :{'M1':7,'M2':6,'M3':3},\n
            'T4' :{'M1':2,'M2':3,'M3':1},\n
            'T5' :{'M1':3,'M2':2,'M3':4},\n
            'T6' :{'M1':1,'M2':8,'M3':7},\n
            'T7' :{'M1':9,'M2':1,'M3':8},\n
            'T8' :{'M1':1,'M2':5,'M3':8},\n
            'T9' :{'M1':8,'M2':2,'M3':9},\n
            'T10':{'M1':6,'M2':1,'M3':7}}
    '''
    def __init__(self, tareas):
        self.tareasBase = tareas
        if len(tareas[list(tareas.keys())[0]])==2:
            self.tareas = tareas
        elif len(tareas[list(tareas.keys())[0]])==3:
            self.tareas={}
            maquinas = list(tareas[list(tareas.keys())[0]].keys())
            for tarea in tareas.keys():
                self.tareas[tarea] = {maquinas[0]+'-'+maquinas[1]:tareas[tarea][maquinas[0]]+tareas[tarea][maquinas[1]],
                maquinas[1]+'-'+maquinas[2]:tareas[tarea][maquinas[1]]+tareas[tarea][maquinas[2]]}
        else:
            raise 'El numero de tareas excede las 3 posibles que soluciona regla de Jhonson'
        self.nombres_tareas = list(self.tareas.keys())
        self.nombres_maquinas = list(list(self.tareas.values())[0].keys())
        print(self.tareas)
        self.EncontrarCombinaciones()
        self.CalcularPosibilidades()
        self.CalcularSecuencias()
        self.CalcularTiemposSecuencias()

    # Encontrar combinaciones posibles segun regla de Jhonson
    def EncontrarCombinaciones(self):
        self.Combinaciones = []
        while self.tareas != {} :
            self.maximo = list(list(self.tareas.values())[0].values())[0]
            for tarea in self.tareas.keys():
                for maquina in self.nombres_maquinas:
                    if self.tareas[tarea][maquina] < self.maximo:
                        self.maximo = self.tareas[tarea][maquina]
            asignacion = []
            for tarea in self.tareas.keys():
                if self.tareas[tarea][self.nombres_maquinas[0]] == self.maximo:
                    asignacion.append([tarea,'I'])
                elif self.tareas[tarea][self.nombres_maquinas[1]] == self.maximo:
                    asignacion.append([tarea,'F'])
            tareas = list(set([tarea[0] for tarea in asignacion]))
            for tarea in tareas:
                self.tareas.pop(tarea)
            self.Combinaciones.append(asignacion)
        return self.Combinaciones

    # Calcular posibles combinaciones segun orden calculado
    def CalcularPosibilidades(self):
        self.SecuenciasPosibles = [[]]
        for combinacion in self.Combinaciones:
            permutaciones = list(permutations(combinacion,len(combinacion)))
            for i in range(len(permutaciones)):
                permutaciones[i] = list(permutaciones[i])
            aux=[]
            for secuancia in self.SecuenciasPosibles:
                for posibilidad in permutaciones:
                    aux.append(secuancia+posibilidad)
            self.SecuenciasPosibles = aux
        return self.SecuenciasPosibles 

    # Calcular cada una de las seceuncias a partir de combinaciones de posibilidades
    def CalcularSecuencias(self):
        self.Secuencias = []
        for secuencia in self.SecuenciasPosibles:
            inicio = []
            fin = []
            for tarea in secuencia:
                if tarea[1]=='F':
                    fin.insert(0, tarea[0])
                else:
                    inicio.append(tarea[0])
            self.Secuencias.append(inicio+fin)
        return self.Secuencias

    # Calcular tiempo de cada combinacion de posibilidad
    def CalcularTiemposSecuencias(self):
        self.TiemposProcesosSecuencias = []
        for secuencia in self.Secuencias:
            self.TiemposProcesosSecuencias.append(self.CalcularTiempoProceso(secuencia))
        return self.TiemposProcesosSecuencias   

    # Calcular tiempo de proceso para cada secuencia
    def CalcularTiempoProceso(self, secuencia):
        duraciones = []
        for tarea in secuencia:
            duraciones.append([j for j in self.tareasBase[tarea].values()])
        matriz = [ [0 for j in i] for i in self.tareasBase.values()]
        for i in range(len(matriz)):
            for j in range(len(matriz[i])):
                if i==0 and j==0:
                    matriz[i][j] = duraciones[i][j]
                elif i==0:
                    matriz[i][j] = matriz[i][j-1] + duraciones[i][j]
                elif j==0:
                    matriz[i][j] = matriz[i-1][j] + duraciones[i][j]
                else:
                    matriz[i][j] = max([matriz[i][j-1],matriz[i-1][j]]) + duraciones[i][j]
        return matriz[i][j]

class SecuenciacionReglaCDS():
    def __init__(self, tareas):
        '''
        SecuenciacionReglaCDS(tareas: dict[str ,dict[str:float, ..., str:float]])\n
        Aplicacion de la regla de CDS, para esto se debe enviar un diccioanrio de diccioanrios
        de la siguiente forma, tareas: diccionario de key:tareas value:diccionarios de maquinas con duraciones 
        de cada tarea respectiva en la maquina, en este caso no todas las tareas deben ser procesadas en el mismo
        orden en a traves de cada una de las maquinas, esta heuristica acepta igual o mayor a 3 maquinas,
        solo que aplicara la misma regla de Jhonson
        Ejemplo:\n
        tareas={'T1' :{'M1':3,'M2':7,'M3':3,'M4':3},\n
                'T2' :{'M1':1,'M2':4,'M3':9,'M4':9},\n
                'T3' :{'M1':7,'M2':6,'M3':3,'M4':1},\n
                'T4' :{'M1':2,'M2':3,'M3':1,'M4':7},\n
                'T5' :{'M1':3,'M2':2,'M3':4,'M4':2},\n
                'T6' :{'M1':1,'M2':8,'M3':7,'M4':9},\n
                'T7' :{'M1':9,'M2':1,'M3':8,'M4':4},\n
                'T8' :{'M1':1,'M2':5,'M3':8,'M4':1},\n
                'T9' :{'M1':8,'M2':2,'M3':9,'M4':4},\n
                'T10':{'M1':6,'M2':1,'M3':7,'M4':4}}
        '''
        self.tareasBase = tareas
        self.tareas = tareas
        self.nombreTareas = list(tareas.keys())
        self.nombreMaquinas = list(list(tareas.values())[0].keys())
        self.diccioanrioTareas = []
        self.Combinaciones = []
        self.SecuenciasPosibles = []
        self.Secuencias = []
        self.TiemposProcesosSecuencias = []
        for i in range(1,len(self.nombreMaquinas)):
            maquinaFicticia1 = self.nombreMaquinas[0:i]
            maquinaFicticia2 = self.nombreMaquinas[-i::]
            tareasAuxiliar={}
            for tarea in self.nombreTareas:
                tareasAuxiliar[tarea] = {
                    "-".join(maquinaFicticia1):sum([self.tareas[tarea][maquina] for maquina in maquinaFicticia1]),
                    "-".join(maquinaFicticia2):sum([self.tareas[tarea][maquina] for maquina in maquinaFicticia2]),}
            self.diccioanrioTareas.append(tareasAuxiliar)
            combinaciones = self.EncontrarCombinaciones(tareasAuxiliar)
            self.Combinaciones.append(combinaciones)
            posibilidades = self.CalcularPosibilidades(combinaciones)
            self.SecuenciasPosibles.append(posibilidades)
            secuencias = self.CalcularSecuencias(posibilidades) 
            self.Secuencias.append(secuencias)
            tiempos = self.CalcularTiemposSecuencias(secuencias)
            self.TiemposProcesosSecuencias.append(tiempos)

    # Encontrar combinaciones posibles segun regla de Jhonson
    def EncontrarCombinaciones(self, tareasAuxiliar):
        Combinaciones = []
        while tareasAuxiliar!= {} :
            nombresMaquinas = list(list(tareasAuxiliar.values())[0].keys())
            maximo = list(list(tareasAuxiliar.values())[0].values())[0]
            for tarea in tareasAuxiliar.keys():
                for maquina in nombresMaquinas:
                    if tareasAuxiliar[tarea][maquina] < maximo:
                        maximo = tareasAuxiliar[tarea][maquina]
            asignacion = []
            for tarea in tareasAuxiliar.keys():
                if tareasAuxiliar[tarea][nombresMaquinas[0]] == maximo:
                    asignacion.append([tarea,'I'])
                elif tareasAuxiliar[tarea][nombresMaquinas[1]] == maximo:
                    asignacion.append([tarea,'F'])
            tareas = list(set([tarea[0] for tarea in asignacion]))
            for tarea in tareas:
                tareasAuxiliar.pop(tarea)
            Combinaciones.append(asignacion)
        return Combinaciones

    # Calcular posibles combinaciones segun orden calculado
    def CalcularPosibilidades(self, combinaciones):
        SecuenciasPosibles = [[]]
        for combinacion in combinaciones:
            permutaciones = list(permutations(combinacion,len(combinacion)))
            for i in range(len(permutaciones)):
                permutaciones[i] = list(permutaciones[i])
            auxiliar=[]
            for secuancia in SecuenciasPosibles:
                for posibilidad in permutaciones:
                    auxiliar.append(secuancia+posibilidad)
            SecuenciasPosibles = auxiliar
        return SecuenciasPosibles

    # Calcular cada una de las seceuncias a partir de combinaciones de posibilidades
    def CalcularSecuencias(self, SecuenciasPosibles):
        Secuencias = []
        for secuencia in SecuenciasPosibles:
            inicio = []
            fin = []
            for tarea in secuencia:
                if tarea[1]=='F':
                    fin.insert(0, tarea[0])
                else:
                    inicio.append(tarea[0])
            Secuencias.append(inicio+fin)
        return Secuencias
    
    # Calcular tiempo de cada combinacion de posibilidad
    def CalcularTiemposSecuencias(self, Secuencias):
        TiemposProcesosSecuencias = []
        for secuencia in Secuencias:
            TiemposProcesosSecuencias.append(self.CalcularTiempoProceso(secuencia))
        return TiemposProcesosSecuencias

    # Calcular tiempo de proceso para cada secuencia
    def CalcularTiempoProceso(self, secuencia):
        duraciones = []
        for tarea in secuencia:
            duraciones.append([j for j in self.tareasBase[tarea].values()])
        matriz = [ [0 for j in i] for i in self.tareasBase.values()]
        for i in range(len(matriz)):
            for j in range(len(matriz[i])):
                if i==0 and j==0:
                    matriz[i][j] = duraciones[i][j]
                elif i==0:
                    matriz[i][j] = matriz[i][j-1] + duraciones[i][j]
                elif j==0:
                    matriz[i][j] = matriz[i-1][j] + duraciones[i][j]
                else:
                    matriz[i][j] = max([matriz[i][j-1],matriz[i-1][j]]) + duraciones[i][j]
        return matriz[i][j]
    
class BranchAndBounds():
    def __init__(self, tareas, verbose: bool = False):
        '''
        tareas={
        'T1' :{'M1':2,'M2':1,'M3':2},
        'T2' :{'M1':1,'M2':5,'M3':1},
        'T3' :{'M1':4,'M2':1,'M3':2},
        'T4' :{'M1':1,'M2':2,'M3':3},
        'T5' :{'M1':3,'M2':1,'M3':1},
        'T6' :{'M1':1,'M2':7,'M3':1},
        'T7' :{'M1':4,'M2':3,'M3':3},
        'T8' :{'M1':3,'M2':2,'M3':4}
        }
        bab = BranchAndBounds(tareas)
        print(bab.CalcularSecuencia(verbose = False))
        print('='*100)
        bab.CrearExcelResultado()
        for key in bab.ecuaciones_bifurcaciones.keys():
            print(key)
            print(bab.ecuaciones_bifurcaciones[key])
        for key in bab.resultado_bifuraciones.keys():
            print(key)
            print(bab.resultado_bifuraciones[key])
        bab.CrearExcelResultado()
        bab.DiagramaGantt()
        '''
        self.tareas = tareas
        self.trabajos = list(self.tareas.keys())
        self.procesos = {trabajo:list(self.tareas[trabajo].keys()) for trabajo in self.tareas.keys()}
        self.duraciones_original = pd.DataFrame(self.tareas).T
        self.duraciones = pd.DataFrame(self.tareas).T
        self.ecuaciones_bifurcaciones = {}
        self.resultado_bifuraciones = {}
        self.secuencia = []

    # Calcular arbol de decisión
    def CalcularSecuencia(self, verbose: bool = False, excel : bool = False):
        while len(self.duraciones.index)>1:
            self.CalcularRamificacionesDisponibles(verbose)
        self.secuencia.append(self.duraciones.index[0])
        if excel:
            pass
        return self.secuencia
    
    def CrearExcelResultado(self, ruta_excel : str = True):
        # Determinar ruta de creación de excel 
        if ruta_excel == True:
            self.ruta_excel_resultado = str(os.getcwd()).replace('\\','/')+'/Resultados Secuenciación Branch and Bound.xlsx'
        else:
            if ruta_excel[-5::] == '.xlsx':
                self.ruta_excel_resultado = ruta_excel
            else:
                raise 'Ruta no valida, no contiene la extension del archivo (<Nombre>.xlsx)'
        # Crear un nuevo libro de Excel
        libro = openpyxl.Workbook()
        hoja = libro.active
        # Variables para mantener el rastreo de la fila actual
        fila_actual = 2
        # Escribir formulación y resultados titulo
        celda = hoja.cell(row=fila_actual, column=2, value='FORMULACIÓN Y RESULTADOS BRANCH AND BOUND')
        celda.font = Font(bold=True, color="0000FF")
        # Calcular longitud
        ancho = len(self.resultado_bifuraciones[list(self.resultado_bifuraciones.keys())[0]].columns)
        # Inicializar contador de bifurcaciones
        bifurcacion = 0
        # Variables para mantener el rastreo de la fila actual
        fila_actual = 4
        for nombre_llave, df in self.resultado_bifuraciones.items():
            bifurcacion += 1
            # Escribir el nombre de la llave en la celda actual
            celda = hoja.cell(row=fila_actual, column=2, value=f'Bifurcacion {bifurcacion}')
            celda.font = Font(bold=True)
            celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
            # Escribir los nombres de las columnas
            for i, col_name in enumerate(df.columns, start=3):
                celda = hoja.cell(row=fila_actual, column=i, value=col_name)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")
            fila_actual += 1
            # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
            for _, row in df.iterrows():
                celda = hoja.cell(row=fila_actual, column=2, value=row.name)
                celda.font = Font(bold=True)
                for i, value in enumerate(row, start=3):
                    celda = hoja.cell(row=fila_actual, column=i, value=value)
                    celda.alignment  = Alignment(horizontal="center")
                fila_actual += 1
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 1
            # Escribir min max texto
            texto = str(df.values).replace('\n','').replace(' [',', max[').replace('[[','min[max[')
            celda = hoja.cell(row=fila_actual, column=2, value='min max')
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            # Escribir min max valor
            celda = hoja.cell(row=fila_actual, column=3, value=df.apply(max, axis=1).min())
            celda.alignment = Alignment(horizontal="center")
            # Escribir min max ecuación
            celda = hoja.cell(row=fila_actual, column=2+ancho+2, value=texto)
            fila_actual += 1
            # Escribir secuenciar
            celda = hoja.cell(row=fila_actual, column=2, value='secuenciar')
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            # Escribir trabajo a secuenciar
            celda = hoja.cell(row=fila_actual, column=3, value=nombre_llave)
            celda.alignment = Alignment(horizontal="center")
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 2
        fila_actual = 4
        for nombre_llave, df in self.ecuaciones_bifurcaciones.items():
            fila_actual += 1
            # Escribir los nombres de las columnas
            for i, col_name in enumerate(df.columns, start=2+ancho+2):
                celda = hoja.cell(row=fila_actual, column=i, value=col_name)
                celda.font = Font(bold=True)
                celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
            # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
            for _, row in df.iterrows():
                for i, value in enumerate(row, start=2+ancho+2):
                    celda = hoja.cell(row=fila_actual, column=i, value=value)
                    celda.alignment  = Alignment(horizontal="center")
                fila_actual += 1
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 4
        # Guardar el libro de Excel
        libro.save(self.ruta_excel_resultado )
        libro.close()

    # Calular ramficacion Branch and Bound faltantes
    def CalcularRamificacionesDisponibles(self, verbose):
        df_ecuacion_bifurcacion = pd.DataFrame()
        df_resultado_bifurcacion = pd.DataFrame()
        # Suposición de cada trabajo como cuello de botella
        for trabajo in self.duraciones.index:
            # Mostrar los resultados por consola
            if verbose:
                print('-'*80)
            # Calculo de fechas de fin de trabajo LBj
            LBj = self.duraciones.cumsum(axis=1).T[[trabajo]].values
            P_t = self.duraciones.drop(trabajo)
            # Calculo de cotas de duración mínima
            for j,proceso in enumerate(P_t.columns):
                matriz_duraciones = P_t.values
                # Tiempos de trabajos no secuenciados en proceso cuello de botella
                duracion_proceso_cuello_botella = matriz_duraciones[:, j]
                # Tiempos de duración despues proceso considerado cuello de botella
                duracion_procesos_restantes = matriz_duraciones[:, j+1:]
                # Conversion de vectores en strings
                str_fin_trabajo_cuello_botella = LBj[j]
                str_tiempos_procesos_no_secuanciados_cuello_botella = str(duracion_proceso_cuello_botella).replace('\n',',').replace(' ','+')
                str_tiempo_despues_proceso_cuello_botella = str(duracion_procesos_restantes).replace('\n',',').replace(', ',',').replace(' ','+')
                str_ecuacion_LBjt = f'{LBj[j]} + sum{str_tiempos_procesos_no_secuanciados_cuello_botella} + min{str_tiempo_despues_proceso_cuello_botella}'
                # Mostrar los resultados por consola
                if verbose:
                    # Salida a texto
                    print(f'   Calculo cotas para trabajo {trabajo} y proceso cuello de botella {proceso}')
                    print('Fin trabajo proceso:', str_fin_trabajo_cuello_botella)
                    print('Tiempos de proceso no secuanciados:', str_tiempos_procesos_no_secuanciados_cuello_botella)
                    print('Duración mínima despues de cuello de botella:', str_tiempo_despues_proceso_cuello_botella)
                    print(f'LB-{trabajo}-{proceso}: {LBj[j]} + sum{str_tiempos_procesos_no_secuanciados_cuello_botella} + min{str_tiempo_despues_proceso_cuello_botella}')
                    print('')
                df_ecuacion_bifurcacion.loc[trabajo, proceso] = str_ecuacion_LBjt
                df_resultado_bifurcacion.loc[trabajo, proceso] = LBj[j][0] + np.sum(duracion_proceso_cuello_botella) + np.min(np.sum(duracion_procesos_restantes, axis = 1))
        trabajo_secuenciado = df_resultado_bifurcacion.max(axis=1).sort_values().head(1).index[0]
        self.secuencia.append(trabajo_secuenciado)
        # Mostrar los resultados por consola
        if verbose:
            print('-'*80)
            print('Resultado Calculos ecuaciones:\n', df_resultado_bifurcacion)
            print('Trabajo secuenciado:', trabajo_secuenciado)
        self.ecuaciones_bifurcaciones[trabajo_secuenciado] = df_ecuacion_bifurcacion
        self.resultado_bifuraciones[trabajo_secuenciado] = df_resultado_bifurcacion
        self.duraciones = self.duraciones.drop(trabajo_secuenciado, axis = 0)

    def CalcularDicciconarioTiempoSecuencia(self, secuencia):
        if secuencia == None:
            secuencia = self.secuencia
        diccionario_tiempos_inicio = {}
        df = self.duraciones_original.copy().reindex(secuencia)
        # Crear una matriz para almacenar los tiempos de inicio de cada tarea
        tiempo_inicio_tareas = pd.DataFrame(index=df.index, columns=df.columns)
        # Calcular el tiempo de inicio para la primera tarea en cada trabajo
        tiempo_inicio_tareas.iloc[0, 0] = df.iloc[0, 0]
        # Calcular el tiempo de inicio para el resto de las tareas en cada trabajo
        for i, trabajo in enumerate(df.index):
            for j, proceso in enumerate(df.columns):
                if i == 0 and j == 0:
                    tiempo_inicio_tareas.iloc[i, j] = 0
                if i == 0 and j != 0:
                    tiempo_inicio_tareas.iloc[i, j] = tiempo_inicio_tareas.iloc[i, j-1] + df.iloc[i, j-1]
                if i != 0 and j == 0:
                    tiempo_inicio_tareas.iloc[i, j] = tiempo_inicio_tareas.iloc[i-1, j] + df.iloc[i-1, j]
                if i != 0 and j != 0:
                    tiempo_inicio_tareas.iloc[i, j] = max(tiempo_inicio_tareas.iloc[i-1, j] + df.iloc[i-1, j], tiempo_inicio_tareas.iloc[i, j-1] + df.iloc[i, j-1])
                diccionario_tiempos_inicio[trabajo+'_'+proceso] = tiempo_inicio_tareas.iloc[i, j]
        return diccionario_tiempos_inicio

    # Generar Diagrama de Gantt
    def DiagramaGantt(self, secuencia: list = None):
        if secuencia == None:
            secuencia = self.secuencia
        diccionario_tiempos_inicio = self.CalcularDicciconarioTiempoSecuencia(secuencia)
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('Maquinas')
        for trabajo in self.trabajos:
            inicios = []
            procesos = []
            duraciones = []
            for nombre_tiempo_inicio in diccionario_tiempos_inicio.keys():
                if trabajo in nombre_tiempo_inicio:
                    inicios.append(diccionario_tiempos_inicio[nombre_tiempo_inicio])
                    trabajo, proceso = nombre_tiempo_inicio.split('_',1)[0], nombre_tiempo_inicio.split('_',1)[1] 
                    duraciones.append(self.tareas[trabajo][proceso])
                    procesos.append(proceso.split('_')[0])
            ax.barh(procesos, duraciones, left=inicios, label=trabajo)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show()

