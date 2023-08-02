from typing import List, Tuple, Dict, Optional, Union
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment
import os

class BranchAndBounds():
    def __init__(self, tareas, verbose: bool = False):
        self.tareas = tareas
        self.trabajos = list(self.tareas.keys())
        self.procesos = {trabajo:list(self.tareas[trabajo].keys()) for trabajo in self.tareas.keys()}
        self.duraciones_original = pd.DataFrame(self.tareas).T
        self.duraciones = pd.DataFrame(self.tareas).T
        self.ecuaciones_bifurcaciones = {}
        self.resultado_bifuraciones = {}
        self.secuencia = []

    # Calcular arbol de decisión
    def CalcularSecuencia(self, verbose: bool = False, excel : bool = False):
        while len(self.duraciones.index)>1:
            self.CalcularRamificacionesDisponibles(verbose)
        self.secuencia.append(self.duraciones.index[0])
        if excel:
            pass
        return self.secuencia
    
    def CrearExcelResultado(self, ruta_excel : str = True):
        # Determinar ruta de creación de excel 
        if ruta_excel == True:
            self.ruta_excel_resultado = str(os.getcwd()).replace('\\','/')+'/Resultados Secuenciación Branch and Bound.xlsx'
        else:
            if ruta_excel[-5::] == '.xlsx':
                self.ruta_excel_resultado = ruta_excel
            else:
                raise 'Ruta no valida, no contiene la extension del archivo (<Nombre>.xlsx)'
        # Crear un nuevo libro de Excel
        libro = openpyxl.Workbook()
        hoja = libro.active
        # Variables para mantener el rastreo de la fila actual
        fila_actual = 2
        # Escribir formulación y resultados titulo
        celda = hoja.cell(row=fila_actual, column=2, value='FORMULACIÓN Y RESULTADOS BRANCH AND BOUND')
        celda.font = Font(bold=True, color="0000FF")
        # Calcular longitud
        ancho = len(self.resultado_bifuraciones[list(self.resultado_bifuraciones.keys())[0]].columns)
        # Inicializar contador de bifurcaciones
        bifurcacion = 0
        # Variables para mantener el rastreo de la fila actual
        fila_actual = 4
        for nombre_llave, df in self.resultado_bifuraciones.items():
            bifurcacion += 1
            # Escribir el nombre de la llave en la celda actual
            celda = hoja.cell(row=fila_actual, column=2, value=f'Bifurcacion {bifurcacion}')
            celda.font = Font(bold=True)
            celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
            # Escribir los nombres de las columnas
            for i, col_name in enumerate(df.columns, start=3):
                celda = hoja.cell(row=fila_actual, column=i, value=col_name)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")
            fila_actual += 1
            # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
            for _, row in df.iterrows():
                celda = hoja.cell(row=fila_actual, column=2, value=row.name)
                celda.font = Font(bold=True)
                for i, value in enumerate(row, start=3):
                    celda = hoja.cell(row=fila_actual, column=i, value=value)
                    celda.alignment  = Alignment(horizontal="center")
                fila_actual += 1
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 1
            # Escribir min max texto
            texto = str(df.values).replace('\n','').replace(' [',', max[').replace('[[','min[max[')
            celda = hoja.cell(row=fila_actual, column=2, value='min max')
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            # Escribir min max valor
            celda = hoja.cell(row=fila_actual, column=3, value=df.apply(max, axis=1).min())
            celda.alignment = Alignment(horizontal="center")
            # Escribir min max ecuación
            celda = hoja.cell(row=fila_actual, column=2+ancho+2, value=texto)
            fila_actual += 1
            # Escribir secuenciar
            celda = hoja.cell(row=fila_actual, column=2, value='secuenciar')
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            # Escribir trabajo a secuenciar
            celda = hoja.cell(row=fila_actual, column=3, value=nombre_llave)
            celda.alignment = Alignment(horizontal="center")
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 2
        fila_actual = 4
        for nombre_llave, df in self.ecuaciones_bifurcaciones.items():
            fila_actual += 1
            # Escribir los nombres de las columnas
            for i, col_name in enumerate(df.columns, start=2+ancho+2):
                celda = hoja.cell(row=fila_actual, column=i, value=col_name)
                celda.font = Font(bold=True)
                celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
            # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
            for _, row in df.iterrows():
                for i, value in enumerate(row, start=2+ancho+2):
                    celda = hoja.cell(row=fila_actual, column=i, value=value)
                    celda.alignment  = Alignment(horizontal="center")
                fila_actual += 1
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 4
        # Guardar el libro de Excel
        libro.save(self.ruta_excel_resultado )
        libro.close()

    # Calular ramficacion Branch and Bound faltantes
    def CalcularRamificacionesDisponibles(self, verbose):
        df_ecuacion_bifurcacion = pd.DataFrame()
        df_resultado_bifurcacion = pd.DataFrame()
        # Suposición de cada trabajo como cuello de botella
        for trabajo in self.duraciones.index:
            # Mostrar los resultados por consola
            if verbose:
                print('-'*80)
            # Calculo de fechas de fin de trabajo LBj
            LBj = self.duraciones.cumsum(axis=1).T[[trabajo]].values
            P_t = self.duraciones.drop(trabajo)
            # Calculo de cotas de duración mínima
            for j,proceso in enumerate(P_t.columns):
                matriz_duraciones = P_t.values
                # Tiempos de trabajos no secuenciados en proceso cuello de botella
                duracion_proceso_cuello_botella = matriz_duraciones[:, j]
                # Tiempos de duración despues proceso considerado cuello de botella
                duracion_procesos_restantes = matriz_duraciones[:, j+1:]
                # Conversion de vectores en strings
                str_fin_trabajo_cuello_botella = LBj[j]
                str_tiempos_procesos_no_secuanciados_cuello_botella = str(duracion_proceso_cuello_botella).replace('\n',',').replace(' ','+')
                str_tiempo_despues_proceso_cuello_botella = str(duracion_procesos_restantes).replace('\n',',').replace(', ',',').replace(' ','+')
                str_ecuacion_LBjt = f'{LBj[j]} + sum{str_tiempos_procesos_no_secuanciados_cuello_botella} + min{str_tiempo_despues_proceso_cuello_botella}'
                # Mostrar los resultados por consola
                if verbose:
                    # Salida a texto
                    print(f'   Calculo cotas para trabajo {trabajo} y proceso cuello de botella {proceso}')
                    print('Fin trabajo proceso:', str_fin_trabajo_cuello_botella)
                    print('Tiempos de proceso no secuanciados:', str_tiempos_procesos_no_secuanciados_cuello_botella)
                    print('Duración mínima despues de cuello de botella:', str_tiempo_despues_proceso_cuello_botella)
                    print(f'LB-{trabajo}-{proceso}: {LBj[j]} + sum{str_tiempos_procesos_no_secuanciados_cuello_botella} + min{str_tiempo_despues_proceso_cuello_botella}')
                    print('')
                df_ecuacion_bifurcacion.loc[trabajo, proceso] = str_ecuacion_LBjt
                df_resultado_bifurcacion.loc[trabajo, proceso] = LBj[j][0] + np.sum(duracion_proceso_cuello_botella) + np.min(np.sum(duracion_procesos_restantes, axis = 1))
        trabajo_secuenciado = df_resultado_bifurcacion.max(axis=1).sort_values().head(1).index[0]
        self.secuencia.append(trabajo_secuenciado)
        # Mostrar los resultados por consola
        if verbose:
            print('-'*80)
            print('Resultado Calculos ecuaciones:\n', df_resultado_bifurcacion)
            print('Trabajo secuenciado:', trabajo_secuenciado)
        self.ecuaciones_bifurcaciones[trabajo_secuenciado] = df_ecuacion_bifurcacion
        self.resultado_bifuraciones[trabajo_secuenciado] = df_resultado_bifurcacion
        self.duraciones = self.duraciones.drop(trabajo_secuenciado, axis = 0)

    def CalcularDicciconarioTiempoSecuencia(self, secuencia):
        if secuencia == None:
            secuencia = self.secuencia
        diccionario_tiempos_inicio = {}
        df = self.duraciones_original.copy().reindex(secuencia)
        # Crear una matriz para almacenar los tiempos de inicio de cada tarea
        tiempo_inicio_tareas = pd.DataFrame(index=df.index, columns=df.columns)
        # Calcular el tiempo de inicio para la primera tarea en cada trabajo
        tiempo_inicio_tareas.iloc[0, 0] = df.iloc[0, 0]
        # Calcular el tiempo de inicio para el resto de las tareas en cada trabajo
        for i, trabajo in enumerate(df.index):
            for j, proceso in enumerate(df.columns):
                if i == 0 and j == 0:
                    tiempo_inicio_tareas.iloc[i, j] = 0
                if i == 0 and j != 0:
                    tiempo_inicio_tareas.iloc[i, j] = tiempo_inicio_tareas.iloc[i, j-1] + df.iloc[i, j-1]
                if i != 0 and j == 0:
                    tiempo_inicio_tareas.iloc[i, j] = tiempo_inicio_tareas.iloc[i-1, j] + df.iloc[i-1, j]
                if i != 0 and j != 0:
                    tiempo_inicio_tareas.iloc[i, j] = max(tiempo_inicio_tareas.iloc[i-1, j] + df.iloc[i-1, j], tiempo_inicio_tareas.iloc[i, j-1] + df.iloc[i, j-1])
                diccionario_tiempos_inicio[trabajo+'_'+proceso] = tiempo_inicio_tareas.iloc[i, j]
        return diccionario_tiempos_inicio

    # Generar Diagrama de Gantt
    def DiagramaGantt(self, secuencia: list = None):
        if secuencia == None:
            secuencia = self.secuencia
        diccionario_tiempos_inicio = self.CalcularDicciconarioTiempoSecuencia(secuencia)
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('Maquinas')
        for trabajo in self.trabajos:
            inicios = []
            procesos = []
            duraciones = []
            for nombre_tiempo_inicio in diccionario_tiempos_inicio.keys():
                if trabajo in nombre_tiempo_inicio:
                    inicios.append(diccionario_tiempos_inicio[nombre_tiempo_inicio])
                    trabajo, proceso = nombre_tiempo_inicio.split('_',1)[0], nombre_tiempo_inicio.split('_',1)[1] 
                    duraciones.append(self.tareas[trabajo][proceso])
                    procesos.append(proceso.split('_')[0])
            ax.barh(procesos, duraciones, left=inicios, label=trabajo)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show()

tareas={
        'T1' :{'M1':2,'M2':1,'M3':2},
        'T2' :{'M1':1,'M2':5,'M3':1},
        'T3' :{'M1':4,'M2':1,'M3':2},
        'T4' :{'M1':1,'M2':2,'M3':3},
        'T5' :{'M1':3,'M2':1,'M3':1},
        'T6' :{'M1':1,'M2':7,'M3':1},
        'T7' :{'M1':4,'M2':3,'M3':3},
        'T8' :{'M1':3,'M2':2,'M3':4}
        }

bab = BranchAndBounds(tareas)
print(bab.CalcularSecuencia(verbose = False))
print('='*100)

bab.CrearExcelResultado()

for key in bab.ecuaciones_bifurcaciones.keys():
    print(key)
    print(bab.ecuaciones_bifurcaciones[key])

for key in bab.resultado_bifuraciones.keys():
    print(key)
    print(bab.resultado_bifuraciones[key])

bab.CrearExcelResultado()
bab.DiagramaGantt()

