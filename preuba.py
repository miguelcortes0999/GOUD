import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd

def save_dict_to_excel(diccionario1, diccionario2, nombre_archivo):
    # Crear un nuevo libro de Excel
    libro = openpyxl.Workbook()
    hoja = libro.active
    # Variables para mantener el rastreo de la fila actual
    fila_actual = 2
    # Escribir formulación y resultados titulo
    celda = hoja.cell(row=fila_actual, column=2, value='FORMULACIÓN Y RESULTADOS BRANCH AND BOUND')
    celda.font = Font(bold=True, color="0000FF")
    # Calcular longitud
    ancho = len(diccionario1[list(diccionario1.keys())[0]].columns)
    # Inicializar contador de bifurcaciones
    bifurcacion = 0
    # Variables para mantener el rastreo de la fila actual
    fila_actual = 4
    for nombre_llave, df in diccionario1.items():
        bifurcacion += 1
        # Escribir el nombre de la llave en la celda actual
        celda = hoja.cell(row=fila_actual, column=2, value=f'Bifurcacion {bifurcacion}')
        celda.font = Font(bold=True)
        celda.alignment  = Alignment(horizontal="center")
        fila_actual += 1
        # Escribir los nombres de las columnas
        for i, col_name in enumerate(df.columns, start=3):
            celda = hoja.cell(row=fila_actual, column=i, value=col_name)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
        fila_actual += 1
        # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
        for _, row in df.iterrows():
            celda = hoja.cell(row=fila_actual, column=2, value=row.name)
            celda.font = Font(bold=True)
            for i, value in enumerate(row, start=3):
                celda = hoja.cell(row=fila_actual, column=i, value=value)
                celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
        # Dejar una fila en blanco después de cada tabla
        fila_actual += 1
        # Escribir min max texto
        texto = str(df.values).replace('\n','').replace(' [',', max[').replace('[[','min[max[')
        celda = hoja.cell(row=fila_actual, column=2, value='min max')
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")
        # Escribir min max valor
        celda = hoja.cell(row=fila_actual, column=3, value=df.apply(max, axis=1).min())
        celda.alignment = Alignment(horizontal="center")
        # Escribir min max ecuación
        celda = hoja.cell(row=fila_actual, column=2+ancho+2, value=texto)
        fila_actual += 1
        # Escribir secuenciar
        celda = hoja.cell(row=fila_actual, column=2, value='secuenciar')
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")
        # Escribir trabajo a secuenciar
        celda = hoja.cell(row=fila_actual, column=3, value=nombre_llave)
        celda.alignment = Alignment(horizontal="center")
        # Dejar una fila en blanco después de cada tabla
        fila_actual += 2
    fila_actual = 4
    for nombre_llave, df in diccionario2.items():
        fila_actual += 1
        # Escribir los nombres de las columnas
        for i, col_name in enumerate(df.columns, start=2+ancho+2):
            celda = hoja.cell(row=fila_actual, column=i, value=col_name)
            celda.font = Font(bold=True)
            celda.alignment  = Alignment(horizontal="center")
        fila_actual += 1
        # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
        for _, row in df.iterrows():
            for i, value in enumerate(row, start=2+ancho+2):
                celda = hoja.cell(row=fila_actual, column=i, value=value)
                celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
        # Dejar una fila en blanco después de cada tabla
        fila_actual += 4
    # Guardar el libro de Excel
    libro.save(nombre_archivo)
    libro.close()

# Ejemplo de uso:
diccionario_ejemplo1 = {
    "tabla1": pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]}),
    "tabla2": pd.DataFrame({'X': [7, 8, 9], 'Y': [10, 11, 12]}),
    "tabla3": pd.DataFrame({'Z': [8, 9, 7], 'L': [1, 11, 82]}),
}

# Ejemplo de uso:
diccionario_ejemplo2 = {
    "tabla1": pd.DataFrame({'A': [10, 20, 30], 'B': [40, 50, 60]}),
    "tabla2": pd.DataFrame({'X': [70, 80, 90], 'Y': [100, 110, 120]}),
    "tabla3": pd.DataFrame({'Z': [170, 180, 190], 'L': [1100, 1110, 1120]}),
}

nombre_archivo = "datos_diccionario.xlsx"
save_dict_to_excel(diccionario_ejemplo1, diccionario_ejemplo2, nombre_archivo)

